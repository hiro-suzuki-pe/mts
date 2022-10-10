## 
# @file loginfo.py 
# @brief 電動バルブの振動データを解析する
# @author ヒロ鈴木
# @date 2021.04.27
#

import sys
import os
import glob
import openpyxl
import numpy as np
from datetime import datetime


##
# @fn get_seqInfo
# @brief data行列のうち，各振動波形の最大値を取得する。
# @param data 
# @retval シーケンス情報
#
def get_seqInfo(data):
    
    (nrow, ncol) = data.shape 
    seqInfo = [ncol]
    if ncol > 2048:
        print('dataError: in get_seqInfo()', data)
        
    for i in range(1,4):
        seqInfo.append(abs(np.max(data[i])))
    return seqInfo

##
# @fn get_logInfo
# @brief シーケンス情報ファイルからログ情報を取得する。
# @param ex_file シーケンス情報ファイル（EXCELファイル）
# @retval ログ情報
#
def get_logInfo(ex_file):
    wb = openpyxl.load_workbook(ex_file)
    ws = wb['シーケンス一覧']
    seqID_l = list(ws.values)

    op_dic = {}
    cl_dic = {}
    for row in seqID_l[1:]:
        seqID = row[1]
#        print(seqID)
        vl = list(wb[seqID].values)
        data = np.array(vl[1:]).T
        if 'OP' in seqID:
            op_dic[seqID] = get_seqInfo(data)
        elif 'CL' in seqID:
            cl_dic[seqID] = get_seqInfo(data)

    logInfo = []
    for seq_dic in [op_dic, cl_dic]:
        if seq_dic == {}:
            logInfo += 4*[0]
        elif len(seq_dic) == 1:
            logInfo += [1] + list(seq_dic.values())[0]
        else:
            log_data = np.array(list(seq_dic.values())).T
            (nrow, ncol) = log_data.shape
            logInfo += [ncol] + list(np.max(log_data, axis=1))
        
    return logInfo


##
# @fn aggregate_files
# @brief Excel化されたログ情報を取得する。
# @retval ログ情報(辞書)
#
def aggregate_files():

    logInfo_dir = {}
    ex_files = glob.glob('*.xlsx')
    
  
    log_dic = {}
    count = 1
    for ex in ex_files:
        print('{:4d} {}\t'.format(count, ex), datetime.now().strftime('%Y/%m/%d %H:%M:%S'))
        log_dic[ex] = get_logInfo(ex)
        count += 1
    return log_dic
    
# *******************************

logdir = 'E:\\VTEX\\ログデータ 鈴木用'
exdir = 'D:\\PROG\\EX'
ofile = 'logFileSummary.xlsx'

cwd = os.getcwd()
os.chdir(exdir)

#ex_file = '2020507-#3-連続380万回-Heリーク後-振動(100).xlsx'
#get_logInfo(ex_file)
#sys.exit(3)
#************************



print('program start:', datetime.now().strftime('%Y/%m/%d %H:%M:%S'))
log_dic = aggregate_files()
os.chdir(cwd)


heading = ['ファイル名', 'OP数', '点数', '振動X', '振動Y', '振動Z',
            'CL数', '点数', '振動X', '振動Y', '振動Z']
wb = openpyxl.Workbook()
ws = wb['Sheet']
for icol in range(len(heading)):
    ws.cell(1,icol+1).value = heading[icol]

irow = 2
for k in log_dic.keys():
#    print('writing', k)
    ws.cell(irow, 1).value = k
    for icol in range(len(log_dic[k])):
        ws.cell(irow, icol+2).value = log_dic[k][icol]
    irow += 1

wb.save(ofile)
print('Program end:', datetime.now().strftime('%Y/%m/%d %H:%M:%S'))


