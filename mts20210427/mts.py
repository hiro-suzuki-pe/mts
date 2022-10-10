## 
# @file mts.py
# @author ヒロ鈴木
# @date 2021.04.27
# @details 
#
import sys
import math
import numpy as np
import openpyxl

## 
# @class m_space
# @brief マハラノビス空間（基準空間)クラス
# @var m_matrix 基準空間（行列）
# @var mean 平均ベクトル
# @var std 標準偏差ベクトル
# @var nx 基準空間の次元数 
# @var ny
# @var data_list  振動データのリスト
# @var data_array 振動データの行列
#
class m_space:
    ##
    # @fn __init__
    # @brief m_spaceクラスのコンストラクタ
    # @param ex_file 基準空間ファイル(Excel）の名称
    # @param m_sheet 基準空間豪列のシート
    # @param av_std_sheet 平均値および平均値の行列
    #
    def __init__(self, data_list=[], ex_file=''):
        
        if ex_file:
            self.load_Matrix()
            return
        
        self.data_list = data_list

        d_l = data_list[1:]
        self.data_array = np.array(d_l, dtype=np.float32)

        (nx,ny) = self.data_array.shape
        self.nx = nx  # Number of samples
        self.ny = ny  # number of parameters
        
        self.mean = np.mean(self.data_array, axis = 0, dtype=np.float32)
        self.std = np.std(self.data_array, axis = 0, ddof=1, dtype=np.float32)
        zero_idx = [i for i in range(self.ny) if self.std[i] == float(0)]
        if zero_idx:
            print('STD error:\n', zero_idx)
            print(self.std)
            sys.exit(2)
            
        norm_array = (self.data_array - self.mean)/self.std           
        co_var = np.cov(norm_array, rowvar=0, bias=0) 
        self.m_matrix = np.linalg.inv(co_var)
        
    ##
    # @fn load_Matrix
    # brief 基準空間(行列)をロードする。
    # @param ex_file 基準空間ファイル(Excel）の名称
    # @param m_sheet 基準空間豪列のシート
    # @param av_std_sheet 平均値および平均値の行列
    #
    def load_Matrix(self, ex_file='Maharanobis_matrix.xlsx', m_sheet = 'Matrix', ave_std_sheet = 'average-stdev' ):

        wb = openpyxl.load_workbook(ex_file)
        ws = wb[m_sheet]
        print('load matrix')

        ml = list(ws.values)
        self.m_matrix = np.array(ml, dtype=float)
        
        ws = wb[ave_std_sheet]
        ml = list(ws.values)
        self.mean = np.array(ml[0], dtype=float)
        self.std = np.array(ml[1], dtype=float)
 
        self.nx = 0
        self.ny = len(self.m_matrix)
        
        self.data_list = []  
        self.data_array = []

    ##
    # @fn M_distance
    # @brief マハラノビス距離を算出する。
    # @param sample 標本データ
    # @retval sampleに対する
    #
    def M_distance(self, sample):

        sample_vec = np.array(sample,dtype=np.float32)
        norm_sample_vec = (sample_vec - self.mean) / self.std
        return np.dot(norm_sample_vec.T, np.dot(self.m_matrix, norm_sample_vec))/self.ny

    ##
    # @fn M_distance
    # @brief SN値を評価するための異常データabnormal
    # @param abnormal 異常データに対するSN値
    # @retval SN値
    #
    def calculate_SN(self, abnormal):
        nd = len(abnormal)
        Ds = [self.M_distance(ab[2:]) for ab in abnormal]
        eta = -10.0 * np.log(sum(1/(d*d) for d in Ds))

        return eta

    ##
    # @fn save_Matrix
    # @brief 基準空間データの保存
    # @param ofile 出力ファイル名（デフォルトは'')
    # @retval SN値
    #
    def save_Matrix(self, ofile=''):
        
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Matrix')
        for i in range(self.ny):
            for j in range(self.ny):
                ws.cell(i+1, j+1).value = self.m_matrix[i][j]
        ws = wb.create_sheet('average-stdev')
        for j in range(self.ny):
            ws.cell(1, j+1).value = self.mean[j]
            ws.cell(2, j+1).value = self.std[j]
        wb.remove(wb['Sheet'])
        wb.save(ofile)            
       
                        
##
# @class mts 
# @brief Mahalanobis-Taguchi システム
# @var normal 正常データ・リスト
# @var abnormal 異常データ・リスト
# @var orth_table 直交表
# data size (shape)
#
class mts:
    ##
    # @fn evaluate_param_effect
    # @brief 直交表を作成する。
    # @param nparm 項目数(パラメータの種類)
    # @retval 直交表
    #
    def evaluate_param_effect(self):
        
        nparm = len(self.normal[0])-2
        m = math.ceil(math.log2(nparm))
        L = self.orthogonal_array(nparm)
        
           
        mah_l = 2**m * [0.0]
        for i in range(2**m):
            (normal, abnormal) = self.extract_data(L[i])
            ms = m_space(normal)
            mah_l[i] = ms.calculate_SN(abnormal)
        
        peff = []
        for i in range(nparm):
            x1 = [mah_l[j] for j in range(2**m) if L[j][i] == 1]
            x2 = [mah_l[j] for j in range(2**m) if L[j][i] == 2]
            peff.append([sum(x1), sum(x2)])

            print('peff[',i,'][0]=', sum(x1), x1)
        return (mah_l, peff)
    

        ##
        # @fn __init__
        # @brief data_list
        # @param data_list データリスト。[0]：データＩＤ，[1]：Y/N， データの採否
        #
        def __init__(self, data_list):
        
        normal = []
        abnormal = []
        
        for d in data_list[1:]:
            if d[1] == 'Y':
                normal.append(d)
            elif d[1] == 'A':
                abnormal.append(d)

        self.normal = normal
        self.abnormal = abnormal
        self.SN = {}
        for s in data_list[1:]:
            self.SN[s[0]] = [0, 0]

#        m = math.ceil(log2(nparm))
#        self.L = self.orthogonal_array(self.nparm)

        print('normal: ', len(normal), ', abnormal: ', len(abnormal),
              ', not used: ', len(data_list) - len(normal) - len(abnormal) - 1)
        
    ##
    # @fn extract_data
    # @brief occurrenceに従って採用項目からなる正常データを抽出する。
    # @param occurrence 項目(パラメータ）の採否ベクトル
    # @retval (normal, abnormal) 正常データおよび異常データ（リストのリスト）
    def extract_data(self, occurrence): 
        nparm = len(self.normal[0]) - 2
        normal = []
        for n in self.normal:
            ne = [n[0]+n[1]] + [n[i+2] for i in range(nparm) if occurrence[i] == 1]
            normal.append(ne)
        
        abnormal = []
        for n in self.abnormal:
            ne = [n[0]+n[1]] + [n[i+2] for i in range(nparm)  if occurrence[i] == 1]
            abnormal.append(ne)

        return (normal, abnormal)

    ##
    # @fn orthogonal_array
    # @brief 直交表を作成する。
    # @param nparm 項目数(パラメータの種類)
    # @retval 直交表
    #
    def orthogonal_array(self, nparm):
        
        m = math.ceil(np.log2(nparm))
        ll = []
        for ip in range(1,2**m):
            a = [(ip // 2**(m-k-1)) % 2 for k in range(m)]
            r = []
            for i in range(2**m):
                b = 0
                for k in range(m):
                    b += a[k] * (i // 2**k)
                r.append(b % 2)
            ll.append(r)
            
#        ll.append((2**k)*[-1.0])
#        print('Orthogonal Array\n', ll)
        return np.array(ll).T + 1

 
