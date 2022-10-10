##
# @file vib.py
# @brief 振動データ処理ファイル
# @author ヒロ鈴木
#

import os
import sys
import shutil
import glob
import openpyxl
import re
import numpy as np

##
# @class vib 
# @brief 振動データクラス
# @var log_dir  ログファイルがあるフォルダ名
# @var cwd 現在のディレクトリ
# @var logfile_list ログファイルのリスト
# @var seqlist_file シーケンス・リストのファイル
# @var parameter_file   パラメータファイル
#
class vib:
    ##
    # @fn    __init__
    # @param log_dir  ログファイルが保存されているフォルダ名
    #
    def __init__(self, log_dir='E:\\VTEX\\ログデータ 鈴木用'):
        self.log_dir = log_dir
        self.cwd = os.getcwd()
        self.logfile_list = 'valve_logfiles.xlsx'
        self.seqlist_file = 'valve_sequences.xlsx'
        self.parameter_file = 'valve_parameters.xlsx'
    
    ##
    # @fn    log_list
    # @brief ログファイルのリストをExcelファイル(*.xlsx) に出力する
    # @param ofile  ログファイルのリストを保存するファイル名
    #
    def log_list(self, ofile=''):
        
        keyword = ['= データ蓄積の設定条件表示', '開始時間:', '動作回数:', 
                   '測定周期(ms):', '位置更新(回):', 'FLASH MODEL:' , 
                   'Data Repeat:', '温度 Min:', '振動 Min:', 
                   '電流 Min:', '定期データ格納', '異常データ格納']
        os.chdir(self.log_dir)
        if not ofile:
            ofile = self.logfile_list
        log_files = glob.glob('*.log')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'FileName'
        ws.cell(row=1, column=2).value = 'Y/N'

        for fno in range(len(log_files)):
            log = log_files[fno]
            with open(log) as lf:
                lines = lf.readlines()
                
            check_flag = 12*[0]
            for l in lines:
                for i in range(len(keyword)):
                    if keyword[i] in l:
                        check_flag[i] = lines.index(l)
        
            ws.cell(row=fno+2, column=1).value = log
            if 0 in check_flag:
                ws.cell(row=fno+2, column=2).value = 'N'
            else:
                ws.cell(row=fno+2, column=2).value = 'Y'
            
        os.chdir(self.cwd)
        wb.save(ofile)
        print(ofile, 'was written')
        
    def sequence_list(self, ifile='', ofile = ''):
        
        if not ifile:
            ifile = self.logfile_list
        if not ofile:
            ofile = self.seqlist_file
        wb = openpyxl.load_workbook(ifile)
        ws = wb.active
        files = list(ws.values)
        wb.close()
        print('<--', ifile)
        
        os.chdir(self.log_dir)
        seq_l = []
        for file in files[1:]:
            if file[1] == 'N':
                continue
            with open(file[0], 'rt') as fi:
                lines = fi.readlines()
            for l in lines:
                if ' << 定期データ蓄積:' in l and '件目のデータ 開始時間' in l and '動作モード:' in l:
                    x = re.findall('\d+', l)
                    dt = '{:04d}{:02d}{:02d}'.format(int(x[1]), int(x[2]), int(x[3]))
                    tm = '{:02d}{:02d}{:02d}'.format(int(x[4]), int(x[5]), int(x[6]))
                    if 'OPEN' in l:
                        ID = 'OP-{}-{}'.format(dt, tm)
                    elif 'CLOSE' in l:
                        ID = 'CL-{}-{}'.format(dt, tm)
                    seq = [file[0], ID, 'Y']
                if 'ユニット名' in l and 'バルブタイプ' in l:
                    x = re.findall('\d+', l)
                    seq.append(int(x[0]))
                if '動作回数' in l and '測定周期' in l and '位置周期' in l:
                    x = re.findall('\d+', l)
                    seq.append(int(x[-2]))
                if '温度' in l and '異常情報' in l and '取得データ数' in l:
                    x = re.findall('\d+', l)
                    seq.append(int(x[-1]))
                    seq_l.append(seq)
                    
        os.chdir(self.cwd)
        wbo = openpyxl.Workbook()
        wso = wbo.worksheets[0]
        
        heading = ['FileName', 'SeqID', 'Y/N', 'valveType', 
                   'si', 'sp']
                   
        for i in range(len(heading)):
            wso.cell(row=1, column=i+1).value = heading[i]
        
        for i in range(len(seq_l)):
            for j in range(len(seq_l[i])):
                wso.cell(row=i+2, column=j+1).value = seq_l[i][j]

        with open('temp.txt', 'wt') as f:
            for s in seq_l:
                print(s, file=f)
        

        wbo.save(ofile)
        print('-->', ofile)
      
    ##
    # @fn get_seqID
    # @brief ログファイルの所定の行から，データのシーケンスIDを作成する
    # @param line データ取得の日時を含むログデータの行
    # @retval シーケンスID
    #
    def get_seqID(self, line):
        x = re.findall('\d+', line)
        dt = '{:04d}{:02d}{:02d}'.format(int(x[1]), int(x[2]), int(x[3]))
        tm = '{:02d}{:02d}{:02d}'.format(int(x[4]), int(x[5]), int(x[6]))
        if 'OPEN' in line:
            ID = 'OP-{}-{}'.format(dt, tm)
        elif 'CLOSE' in line:
            ID = 'CL-{}-{}'.format(dt, tm)
        else:
            ID = 'OT-{}-{}'.format(dt, tm)
        return ID
 
    ##
    # @fn parameter_file_heading
    # @brief パラメータファイルの見出し行を書き出す
    # @retval 
    #
    def parameter_file_heading(self):
        h = ['seqID', 'Y/N', 'type(0:pully,1:coupling)', 'samplingPoint']
        for ax in['x', 'y', 'z', 'Tq']:
            for i in range(11):
                h.append('{}({})ave'.format(ax, i))
                h.append('{}({})dispersion'.format(ax, i))
                h.append('{}({})(max-min)/2'.format(ax, i))
        return h
    
    ##
    # @fn calc_parameter
    # @brief MTSに掛ける特性パラメータを算出する
    # @retval 
    #
    def calc_parameter(self, xyz):
        
        ar = np.array(xyz).T
        sp = len(xyz)
        xl = []
        for i in range(1,5):
            for j in range(0, 1100, 100):
                if j < sp:
                    n = min(sp-j, 100)
                    xl.append( np.mean(ar[i][j:j+n]))
                    if (sp % 100) == 1:
                        xl.append( np.var(ar[i][j:j+n],ddof=0))
                    else:
                        xl.append( np.var(ar[i][j:j+n],ddof=1))
                    xl.append( (np.max(ar[i][j:j+n]) - np.min(ar[i][j:j+n]))/2)
                else:       
                    xl += [0.0, 0.0, 0.0]
        return xl

    ##
    # @fn extract_parameters
    # @brief MTSに掛ける特性パラメータを算出する
    # @retval 
    #
    def extract_parameters(self, ifile='', ofile = ''):
        
        if not ifile:
            ifile = self.logfile_list
        if not ofile:
            ofile = self.parameter_file
        wb = openpyxl.load_workbook(ifile)
        ws = wb.active
        files = list(ws.values)
        wb.close()
        print('<--', ifile)
        
        os.chdir(self.log_dir)
        parm_dic = {}
        parm_dic['OP'] = []
        parm_dic['CL'] = []
        for file in files[1:]:
            e = file[0].split('-')
            if len(e) >= 2 and e[1] in ['#7' , '#9', '#10']:
                mtype = 0
            else:
                mtype = 1
                
            if file[1] == 'N':
                continue
            with open(file[0], 'rt') as fi:
                lines = fi.readlines()
            
            in_data = False
            for l in lines:
                if ' << 定期データ蓄積:' in l and '件目のデータ 開始時間' in l and '動作モード:' in l:
                    ID = self.get_seqID(l)
                    parm = [ID, 'Y', mtype]
                    xyz = []
                    in_data = True
                elif '温度(動作前' in l and '動作後' in l and '異常情報' in l and '取得データ数' in l:
                    x = re.findall('\d+', l)
                    sp = int(x[-1])
                    parm.append(sp)
                elif re.match('\s*\d+\s*(,\s*-?\d+)+\s*', l):
                    x = re.split(',', l)
                    if int(x[0]) < 1100:
                        xyz.append([int(xx) for xx in x])                    
                elif re.match('^\s*$', l) and in_data:
                    in_data = False
                    for i in range(sp, 1100):
                        xyz.append([0, 0, 0, 0, 0, 0])
                    parm += self.calc_parameter(xyz)
                    if not 'OT-' in parm[0]:
                        parm_dic[parm[0][:2]].append(parm)


        os.chdir(self.cwd)
        wb = openpyxl.Workbook()
        wb.create_sheet('OPEN')
        wb.create_sheet('CLOSE')
        wb.remove(wb.worksheets[0])
        
        for sheet in ['OPEN', 'CLOSE']:
            ws = wb[sheet]
            sh = sheet[:2]
            heading = self.parameter_file_heading()
            for i in range(len(heading)):
                ws.cell(row=1, column=i+1).value = heading[i]
            
            for i in range(len(parm_dic[sh])):
                for j in range(len(heading)):
                    ws.cell(row=i+2, column=j+1).value = parm_dic[sh][i][j]
        wb.save(ofile)                   
        print('-->', ofile)
         
    ##
    # @fn load_parameters
    # @brief 特性パラメータをファイルからロードする。
    # @param sname 特性パラメータのワークシート
    # @param unused_parameter 使用しない特性パラメータの採否フラグベクトル
    # @param ifile  特性パラメータファイル（*.xlsx)
    # @retval 特性パラメータからなるテーブル
    #
    def load_parameters(self, sname, unused_parm=[], ifile=''):
        
        if not ifile:
            ifile = self.parameter_file

        table = []
        wb = openpyxl.load_workbook(ifile)
        ws = wb[sname]
        print('<--', ifile)
        
        
        for irow in range(1,ws.max_row):
            row_l = []
            for icol in range(ws.max_column):
                if not icol-2 in unused_parm:
                    row_l.append(ws.cell(irow+1, icol+1).value)
            table.append(row_l)
        return table            
