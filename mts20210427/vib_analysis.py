## 
# @file vib_analysis.py 
# @brief 電動バルブの振動データを解析する
# @author ヒロ鈴木
# @date 2021.04.27
#



import sys
from datetime import datetime
import openpyxl
import math
import mts
import vib
import matplotlib.pyplot as plt

print('Program arguments:', sys.argv)


vib = vib.vib(log_dir='d:\\Prog\\data')
              
stage = {'log list': False,
         'Sequence list': False,
         'Extract parameters': True,
         'Build MTS': False,
         'Evaluate parameters': False}
         
print('Welcome to Vibration Analyzer: ', datetime.now())
print('\n')

for k in stage.keys():
    
    if stage[k]:  
        print(k, ': execute', datetime.now())
    else:
        print(k, ': skipped')
        continue
    
    if k == 'log list':
        vib.log_list()
        
    if k == 'Sequence list':
        vib.sequence_list()
    
    if k == 'Extract parameters':
        vib.extract_parameters()

    if k == 'Build MTS':
        unused_parm = [0, 35, 36, 37, 71, 72, 73, 107, 108, 109, 143, 144, 145]
        samples = vib.load_parameters('CLOSE', unused_parm)
        samples_wo_ID = [s[2:] for s in samples]
        print(len(samples), 'samples loaded. (', len(samples[0]),'parameters)')
        try:
            ms_close = mts.m_space(samples_wo_ID)
            print('m_space: ', datetime.now())
        except:
            print('exception @m_space(). samples:\n', samples[0])
        
        M_of_samples = {}
        count  = 0
        for s in samples:
            M_of_samples[s[0]] = ms_close.M_distance(s[2:])
            count += 1
        print('sample count: ', count)
        
        freq = [0 for i in range(51)]
        for k in M_of_samples.keys():
            idx = min(int(10.0 * M_of_samples[k]), 50)
            print('{},{:2d},{:6.2f}'.format(k, idx, M_of_samples[k]))
            freq[idx] += 1
        
        print('Occurrance')
        for i in range(50):
            print('{:6.2f}-{:6.2f}: {:2d}'.format(0.1*i, 0.1*(i+1), freq[i]))
        print('{:6.2f}-      : {:2d}'.format(5.0, freq[i]))
        
        ma = [v for v in M_of_samples.values()]
        print('Average:', sum(ma)/len(ma))
        print('min:', min(ma))
        print('max:', max(ma))
        
        x = [0.1 * i for i in range(51)]
        
        plt.plot(x,freq)
        plt.show()
        
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.cell(1,1).value = 'seqID'
        ws.cell(1,2).value = 'M'

        irow = 2
        for k in M_of_samples.keys():
            ws.cell(irow,1).value = k
            ws.cell(irow,2).value = M_of_samples[k]
            irow += 1
        wb.save('M.xlsx')
        print('-->', 'M.xlsx')
        
        ms_close.save_Matrix('Maharanobis_matrix.xlsx')
        
    if k =='Evaluate parameters':
        samples = vib.load_parameters('CLOSE', unused_parm)
        mt = mts.mts(samples)
        if not mt.abnormal:
            print('Set some abnormal data in sample[]')
            sys.exit(1)
        (mah_l, sn) = mt.evaluate_param_effect()
        
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = 'SN for Row'
        ws.cell(1,1).value = 'No'
        ws.cell(1,+2).value = 'M'
        nparm = len(samples[0])-2
        for i in range(nparm):
            ws.cell(1,i+3).value = samples[0][i+2]
        
        L = mt.orthogonal_array(nparm)
        m = math.ceil(math.log2(nparm))
        for i in range(2**m):
            ws.cell(i+2,1).value = i+1
            ws.cell(i+2,2).value = mah_l[i]
            for j in range(nparm):
                ws.cell(i+2,j+3).value = L[i][j]
                
        ws = wb.create_sheet('SN for Parm')
        ws.cell(1,1).value = 'Level'
        for i in range(nparm):
            ws.cell(1,i+2).value = samples[0][i+2]

        for i in range(2):
            ws.cell(i+2,1).value = i+1
            for j in range(nparm):
                ws.cell(i+2,j+2).value = sn[j][i]
       
        wb.save('valve_result.xlsx')
        print('-->', 'valve_result.xlsx')
        
    #==================
    print('All done', datetime.now())
    
    
